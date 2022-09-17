import datetime
from pathlib import Path

import openpyxl.styles.numbers
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from constants import EXCEL_FILE_NAME, NAME_COLUMN
from setup import BASE_DIR
from utils.model import Report

PATH_TO_FILE = BASE_DIR.joinpath('media/' + EXCEL_FILE_NAME)


def close_workbook(wb: Workbook):
    """
    Close workbook
    :param wb: Workbook object
    :type wb: Workbook
    :return:
    :rtype:
    """
    wb.close()


def save_file(wb: Workbook, path_to_save_file: Path = PATH_TO_FILE):
    """
    Saving the workbook to a file
    :param wb: Workbook object
    :type wb: Workbook
    :param path_to_save_file: Path to save file
    :type path_to_save_file: Path
    :return:
    :rtype:
    """
    path_to_save_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path_to_save_file)


def add_title_report(sheet: Worksheet, column_index: int = 0) -> Worksheet:
    """
    Adding a title to a report
    :param sheet: Worksheet object
    :type sheet: Worksheet
    :param column_index: Column index to add cell header
    :type column_index: int
    :return: Update worksheet object with title column
    :rtype: Worksheet
    """
    shift = 0
    for val in NAME_COLUMN.values():
        sheet.cell(column=column_index + shift, row=1, value=val)
        sheet.cell(column=column_index + shift, row=1).alignment = Alignment(vertical='center', horizontal='center')
        shift += 1
    return sheet


def formatting_column_by_type(columns: list, ws: Worksheet, format_pattern: str) -> Worksheet:
    """
    Formatting a Column in custom format
    :param columns: List of column name letters
    :type columns: list
    :param ws: Worksheet object
    :type ws: Worksheet
    :param format_pattern: Format pattern
    :type format_pattern: str
    :return: Update Worksheet object
    :rtype: Worksheet
    """
    for letter in columns:
        column_letter = ws[letter]
        for i in range(len(column_letter)):
            column_letter[i].number_format = format_pattern
    return ws


def settings_width_column(ws: Worksheet, col_width: list):
    """
    Setting the column width
    :param ws: Worksheet object
    :type ws: Worksheet
    :param col_width: List of sheet column widths
    :type col_width: list
    :return:
    :rtype:
    """
    for i in range(len(col_width)):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = col_width[i] + 2
        if col_letter in ['B', 'E', 'G']:
            ws.column_dimensions[col_letter].width = col_width[i] + 3 + int(col_width[i] // 3)


def get_max_size_width_column(ws: Worksheet) -> list:
    """
    Getting the maximum character width of columns
    :param ws: Worksheet object
    :type ws: Worksheet
    :return: List of sheet column widths
    :rtype: list
    """
    i = 0
    col_width = []
    for col in ws.columns:
        for j in range(len(col)):
            if j == 0:
                col_width.append(len(str(col[j].value)))
            else:
                if col_width[i] < len(str(col[j].value)):
                    col_width[i] = len(str(col[j].value))
        i = i + 1
    return col_width


def formatting_column(ws: Worksheet) -> Worksheet:
    """
    Column formatting
    :param ws: Worksheet object
    :type ws: Worksheet
    :return: Update Worksheet object
    :rtype: Worksheet
    """
    ws = add_title_report(ws, 1)
    ws = add_title_report(ws, 4)
    ws.cell(row=1, column=7, value='Результат').alignment = Alignment(vertical='center', horizontal='center')
    ws = formatting_column_by_type(['A', 'D'], ws, openpyxl.styles.numbers.BUILTIN_FORMATS[15])
    ws = formatting_column_by_type(['C', 'F'], ws, openpyxl.styles.numbers.BUILTIN_FORMATS[20])
    ws = formatting_column_by_type(['B', 'E', 'G'], ws, openpyxl.styles.numbers.BUILTIN_FORMATS[44])
    return ws


def add_report_data(row_start_index: int, column_start_index: int, data: list[Report], ws: Worksheet) -> Worksheet:
    """
    Adding data to a worksheet
    :param row_start_index:
    :type row_start_index: int
    :param column_start_index:
    :type column_start_index: int
    :param data: List report data
    :type data: list[Report]
    :param ws: Worksheet object
    :type ws: Worksheet
    :return: Worksheet object
    :rtype: Worksheet
    """
    for i in range(len(data)):
        ws.cell(row=row_start_index + i, column=column_start_index).value = data[i].date.strftime('%d.%m.%Y')
        ws.cell(row=row_start_index + i, column=column_start_index + 1, value=data[i].exchange_rate)
        ws.cell(row=row_start_index + i, column=column_start_index + 2).value = data[i].time.strftime('%H:%M')
    return ws


def add_division_formula(cnt_element: int, ws: Worksheet) -> Worksheet:
    """
    Adding a division formula
    :param cnt_element: Number of elements to add formula
    :type cnt_element: int
    :param ws: Worksheet object
    :type ws: Worksheet
    :return: Worksheet object
    :rtype: Worksheet
    """
    for i in range(1, cnt_element + 1):
        ws[f'G{i + 1}'] = f'=$B{i + 1}/$E{i + 1}'
    ws[f'H1'] = f'=SUM(B:B)'
    ws[f'H2'] = f'=SUM(E:E)'
    return ws


def generate_report(data: dict):
    """
    Report generation
    :param data: Report data
    :type data: dict
    :return:
    :rtype:
    """
    wb = Workbook()
    ws = wb.active
    shift = 1
    for el in data.values():
        add_report_data(2, shift, el, ws)
        shift += 3
    add_division_formula(min(len(data['USD_RUB']), len(data['JPY_RUB'])), ws)
    ws = formatting_column(ws)
    settings_width_column(ws, get_max_size_width_column(ws))

    save_file(wb)
    close_workbook(wb)


if __name__ == '__main__':
    d = datetime.datetime.now()
    generate_report({
        'USD_RUB': [Report(**{"date": d.date(), "exchange_rate": 123.3, "time": d.time()})],
        'JPY_RUB': [Report(**{"date": d.date(), "exchange_rate": 0.3, "time": d.time()})]
    })
